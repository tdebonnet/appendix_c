#!/usr/bin/env python
# coding: utf-8

# # Ecriture de l'annexe C
# Resources.

# In[1]:


import csv
import pandas as pd
import numpy as np
import pickle
import os
import sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
# In[ ]:


#File path from the interface
def load_file_path():
    try:
        with open('file_path.pickle', 'rb') as file:
            file_path = pickle.load(file)
            return file_path
    except FileNotFoundError:
        return None

# Utilisation de la fonction pour récupérer le chemin du fichier
file_path = load_file_path()

#Save path 

def load_save_path():
    try:
        with open('save_path.pickle', 'rb') as file:
            save_path = pickle.load(file)
            return save_path
    except FileNotFoundError:
        return None

def load_file_name():
    global file_name_saved  # Ajoutez cette ligne pour indiquer que vous utilisez la variable globale
    try:
        with open('file_name_saved.pickle', 'rb') as file:
            file_name_saved = pickle.load(file)
            if not file_name_saved.endswith('.xlsx'):
                file_name_saved += '.xlsx'
            return file_name_saved
    except FileNotFoundError:
        return None
        
file_name_saved = load_file_name()
#print(file_name_saved)
# Utilisation des fonctions pour récupérer les informations nécessaires
#file_path = load_file_path()
save_path = load_save_path()


# In[99]:


#Reading and copy of the relevant elements of the csv.
data_list=[]
test_products = False
empty_list=[""]*9



with open(file_path, encoding = "ISO-8859-1") as file_name:
    file_read = csv.reader(file_name, delimiter=";")
    next_line = None
    for row in file_read:
        cleaned_row = [cell.strip() for cell in row]
        if not any(cleaned_row):
            # If the line is empty, go to the next line
            continue
       
        #add products
        if "Products" in cleaned_row:
            test_products = True
            data_list.append(empty_list)
            data_list.append(cleaned_row)
            #print(cleaned_row)
            
        elif "End" in cleaned_row:
            test_products = False
        elif test_products:
            # If we are between "Products" and "Avoided products", add the line to the list
            data_list.append(cleaned_row)
        


# In[100]:


# Exclude empty list and the name associated
exclude_words = ["Products", "Avoided products", "Resources", "Materials/fuels", "Electricity/heat", "Emissions to air",
                 "Emissions to water", "Emissions to soil", "Final waste flows", "Non material emissions",
                 "Waste to treatment","Social issues","Economic issues", "Input parameters", "Calculated parameters", "End"]

# Delete empty lists 
result_list = []
exclude_flag = False
i=-1

for i in range(0,len(data_list)-1):
    current_name = data_list[i][0]
    if  current_name not in exclude_words:
        result_list.append(data_list[i])
    else:  
        indice=exclude_words.index(current_name)
        if not data_list[i+1][0] in exclude_words and len(data_list[i+1][0])>0 :
            result_list.append(data_list[i])


# In[101]:


annex_C=pd.DataFrame(result_list)#data_list)
annex_C = annex_C.drop(columns=annex_C.columns[-1]) #remove the UUID
#annex_C


# In[102]:


#Remove ecoinvent process modified by CIRAIG (starting with z.)

# Find index of "products" lines
product_indices = annex_C[annex_C[annex_C.columns[0]] == 'Products'].index
CIRAIG_ecoinvent_index = [index for index, value in enumerate(annex_C.iloc[:, 0]) if str(value).startswith('z.') or str(value).startswith('Z.')]
rows_to_remove_1=[]
rows_to_remove_2=[]
i=0
#for n in product_indices:
for i in range (0,len(product_indices)-1):
    n=product_indices[i]
    #print(n)
    if n + 1 in CIRAIG_ecoinvent_index:
        #print("yes")
        rows_to_remove_1.append(product_indices[i])
        #print(product_indices[i])
        rows_to_remove_2.append(product_indices[i+1]-1)
if product_indices[-1]<CIRAIG_ecoinvent_index[-1]:
    rows_to_remove_1.append(product_indices[-1])
    rows_to_remove_2.append(len(annex_C)-1)
        #print(product_indices[i])
rows_to_remove=[]
#print(CIRAIG_ecoinvent_index)
#print(rows_to_remove_1)
#print(rows_to_remove_2)
for start, end in zip(rows_to_remove_1, rows_to_remove_2):
    sublist = list(range(start, end + 1))
    
    rows_to_remove.extend(sublist)
annex_C.drop(index=rows_to_remove, inplace=True)
annex_C.reset_index(drop=True, inplace=True)


# In[103]:


#Add the location in a new column
pattern = r'\{(.*?)\}'
annex_C['Location'] = annex_C[annex_C.columns[0]].str.extract(pattern)

# Remplace values different than Location by ''
annex_C['Location'].fillna('', inplace=True)

# If there is no Location --> ''
annex_C.loc[annex_C[annex_C.columns[0]].str.contains(pattern) == False, 'Location'] = ''
# Put the column Location in second position
annex_C.insert(3, 'Location', annex_C.pop('Location'))


#################Modify ressources line to align them with the other columns.
modified_subsets = []
# Identify indices of each biosphere flows.
start_indices = []
end_indices=[]
test_biosphere=False
added_index = False
for i, row in enumerate(annex_C.itertuples(index=False)):  # Use itertuples for faster iteration
    if row[0] == "Emissions to air" or row[0] == "Emissions to soil" or row[0] == "Emissions to water":
        start_indices.append(i)
        test_biosphere = True
    elif test_biosphere and("Input parameters" in row or "Calculated parameters" in row or len(row[0]) == 0):
        end_indices.append(i)
        test_biosphere = False
        added_index=True
    
# If no index was added in the conditions above, add the index of the last row
if not added_index:
    end_indices.append(len(annex_C) - 1)      
new_column_order = [0,2,4,1,5,6,7,8,3]  # Index des colonnes dans l'ordre souhaité

# Parcourez les indices de début pour créer les sous-ensembles modifiés
for start_idx, end_idx in zip(start_indices, end_indices):
    subset = annex_C.iloc[start_idx + 1: end_idx]
    modified_subset = subset.iloc[:, new_column_order]
    modified_subsets.append(modified_subset)
    
# Remplacez les sous-ensembles modifiés dans le DataFrame original
for start_idx, modified_subset in zip(start_indices, modified_subsets):
    annex_C.iloc[start_idx + 1: start_idx + 1 + len(modified_subset)] = modified_subset.values

# Réinitialisez les index pour le DataFrame
column_names_annex_C = annex_C.columns

#################Modify parameters line to align them with the other columns.
modified_subsets_parameters = []
# Identify indices of each parameters.
start_indices= []
end_indices=[]
test_param=False
added_index = False
col3_index = 3#'Location'  # Indice de la colonne 2
col4_index = 4  # Indice de la colonne 3
col5_index = 5 
col6_index = 6 
for i, row in enumerate(annex_C.itertuples(index=False)):  # Use itertuples for faster iteration
    if row[0] == "Input parameters" or row[0] == "Calculated parameters":
        start_indices.append(i)
        test_param = True
    elif test_param and("End" in row or len(row[0]) == 0):
        end_indices.append(i)
        # Construct the concatenated string
        concatenated_values = (
            str(annex_C.iat[i, col3_index]) + ';' +
            str(annex_C.iat[i, col4_index]) + ';' +
            str(annex_C.iat[i, col5_index]) + ';' +
            str(annex_C.iat[i, col6_index])
        )
        
        # Assign the concatenated string to the cell
        annex_C.iat[i, col3_index] = concatenated_values
        
        test_param = False
        added_index=True
    
# If no index was added in the conditions above, add the index of the last row
if not added_index:
    end_indices.append(len(annex_C) - 1)      
new_column_order = [0,1,2,3,4,5,6,7,8]#2,4,1,5,6,7,8,3]  # Index des colonnes dans l'ordre souhaité

# Parcourez les indices de début pour créer les sous-ensembles modifiés
for start_idx, end_idx in zip(start_indices, end_indices):
    subset = annex_C.iloc[start_idx + 1: end_idx]
    modified_subset = subset.iloc[:, new_column_order]
    modified_subsets_parameters.append(modified_subset)
    
# Remplacez les sous-ensembles modifiés dans le DataFrame original
for start_idx, modified_subset in zip(start_indices, modified_subsets_parameters):
    annex_C.iloc[start_idx + 1: start_idx + 1 + len(modified_subset)] = modified_subset.values

# Réinitialisez les index pour le DataFrame
column_names_annex_C = annex_C.columns

##########Delete uncertainties values and distribution. 
# indices to drop 
columns_to_drop = [ 4, 5, 6,7]  # Indices des colonnes à supprimer
# Supprimer les colonnes spécifiées
annex_C = annex_C.drop(columns=annex_C.columns[columns_to_drop])



##################### Mettre les titres 

# Parcourez les colonnes et mettez à jour les valeurs si "Products" est présent dans la première colonne
#annex_C = annex_C.reset_index(drop=True)
product_indices = annex_C[annex_C[annex_C.columns[0]] == 'Products'].index
for i in range (0,len(product_indices)):
    n=product_indices[i]
    annex_C[0][n]="name"
    annex_C[1][n]="unit"
    annex_C[2][n]="amount"
    annex_C['Location'][n]="location"
    annex_C[7][n]="description"
    
input_param_indices = annex_C[annex_C[annex_C.columns[0]] == 'Input parameters'].index
for i in range (0,len(input_param_indices)):
    n=input_param_indices[i]
    annex_C[0][n]="Input parameters"
    annex_C[1][n]="unit"
    annex_C[2][n]="distributionType"
    annex_C['Location'][n]="distribution"
    annex_C[7][n]="description"
    
calc_param_indices = annex_C[annex_C[annex_C.columns[0]] == 'Calculated parameters'].index
for i in range (0,len(calc_param_indices)):
    n=calc_param_indices[i]
    annex_C[0][n]="Calculated parameters"
    annex_C[1][n]="unit"
    annex_C[2][n]="distributionType"
    annex_C['Location'][n]="distribution"
    annex_C[7][n]="description"
    
annex_C.replace(";;;", "", inplace=True)    
###########Parameters

# In[104]:


#Add parameters in a new Excel sheet
#Reading and copy of the relevant elements of the csv.
parameter_list=[]
test_project_input_parameters=False
test_project_calculated_parameters=False
empty_list=[""]*9

with open(file_path, encoding = "ISO-8859-1") as file_name:
    file_read = csv.reader(file_name, delimiter=";")
    next_line = None
    for row in file_read:
        cleaned_row = [cell.strip() for cell in row]
        if not any(cleaned_row):
            # If the line is empty, go to the next line
            continue            
        #add avoided products
        if "Project Input parameters" in cleaned_row:
            parameter_list.append(cleaned_row)
            #print("yes")
            test_project_input_parameters = True
        elif "End" in cleaned_row:
            test_project_input_parameters = False
        elif test_project_input_parameters:
            # If we are between "Avoided products" and "resources", add the line to the list
            parameter_list.append(cleaned_row)
            
        #add avoided products
        if "Project Calculated parameters" in cleaned_row:
            parameter_list.append(cleaned_row)
            test_project_calculated_parameters = True
        elif "End" in cleaned_row:
            test_project_calculated_parameters = False
        elif test_project_calculated_parameters:
            parameter_list.append(cleaned_row)


# In[105]:


exclude_words_parameters= ["Project Input parameters", "End", "Project  Calculated parameters", "End"]

# Delete empty lists 
result_list_parameter = []
exclude_flag = False
i=-1

for i in range(0,len(parameter_list)-1):
    current_name = parameter_list[i][0]
    if current_name not in exclude_words_parameters:
        result_list_parameter.append(parameter_list[i])
    else:  
        indice=exclude_words_parameters.index(current_name)
        if not parameter_list[i+1][0] in exclude_words or  len (parameter_list[i+1][0])>0:
            result_list_parameter.append(parameter_list[i])
        #Delete calculated parameters lines when they are empty


# In[106]:


df_parameter=pd.DataFrame(result_list_parameter)
column_names=['name', 'value', 'distribution type', '', '', '', '', 'description']
if df_parameter.shape[1] == len(column_names) + 1:
    name_new_column = ''  # Remplacez par le nom que vous voulez attribuer
    
    # Ajouter la nouvelle colonne au DataFrame
    #df[name_new_column] = None  # Vous pouvez remplacer 'None' par des valeurs par défaut si nécessaire
    column_names.append(name_new_column)
df_parameter.columns = column_names
#df_parameter


# # Add on Excel

# In[107]:


from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import IllegalCharacterError

# Creation of new Workbook
wb = Workbook()

# Activate the sheet
ws = wb.active
ws.title = "System description"

# Description project
text_lines = ['Annex C : Projet']#, 'Projet']

# Add lines to Excel file
for idx, text in enumerate(text_lines, start=1):
    cell = ws.cell(row=idx, column=1, value=text)

# Write dataframe and delete first row
for r_idx, row in enumerate(dataframe_to_rows(annex_C, index=False), start=len(text_lines)+1):
    for c_idx, value in enumerate(row, start=1):
        
        #print(value)
        #ws.cell(row=r_idx, column=c_idx, value=value)
        
        try:
            ws.cell(row=r_idx, column=c_idx, value=value)
        except IllegalCharacterError as e:
            # Handle the exception here, e.g., replace the problematic characters or skip the cell
            print(f"Exception: {e}. Cell content: {value}")
ws.delete_rows(2)

for row in ws.iter_rows(min_row=len(text_lines) + 2, max_row=ws.max_row):#, min_col=1, max_col=1):
    if row[0].value in ["Products", "Avoided products", "Resources", "Materials/fuels", "Electricity/heat", "Emissions to air",
             "Emissions to water", "Emissions to soil", "Final waste flows", "Non material emissions",
             "Waste to treatment", "Input parameters", "Calculated parameters"]:
        # Obtient la cellule correspondante dans la première colonne
        for cell in row:
            #cell = ws.cell(row=ligne_actuelle, column=1)
        # Applique le style de police en gras
            cell.font = Font(bold=True)

# Put in bold the name line
for row in ws.iter_rows(min_row=len(text_lines) + 2, max_row=ws.max_row):#, min_col=1, max_col=1):
    if row[0].value == "name" or row[0].value == "Input parameters"or row[0].value == "Calculated parameters":
        for cell in row:
            cell.font = Font(bold=True)
    #        cell.fill = PatternFill(start_color="00A1C0", end_color="00A1C0", fill_type="solid")
# Define blue color
fill = PatternFill(start_color='00A1C0', end_color='00A1C0', fill_type='solid')
# Définir le style de police avec la couleur de texte blanche
white_font = Font(color="FFFFFF")

# Flag to indicate if the next cell should be formatted

blue_next_row = False

# Parcourir les lignes de la feuille de calcul
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if blue_next_row:
        # Mettre en bleu la ligne courante
        for cell in row:
            cell.fill = fill
            cell.font = white_font
            #print("ok")
        # Réinitialiser le drapeau après avoir mis en bleu la ligne
        blue_next_row = False
    
    if any(cell.value == "name" for cell in row if cell.value is not None):
        # Activer le drapeau pour mettre en bleu la ligne suivante
        blue_next_row = True
        
cell = ws.cell(row=1, column=1)
cell.value = "Annex C : Projet"
cell.font = Font(bold=True, size=16)

#adjust width column

import openpyxl

# Définir la largeur des colonnes en pixels
col_width_col1 = 70  # Largeur de la première colonne en pixels
col_width_others = 15  # Largeur des autres colonnes en pixels

# Définir la largeur de la première colonne (indice 1)
ws.column_dimensions[ws.cell(row=1, column=1).column_letter].width = col_width_col1

# Définir la largeur des autres colonnes (à partir de l'indice 2)
for col_idx, col in enumerate(ws.columns, start=1):
    if col_idx != 1:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = col_width_others

#Add the parameter tab

# New Sheet "Parameters"
ws_parameters = wb.create_sheet(title="Parameters")

for idx, text in enumerate(text_lines, start=1):
    cell_bis = ws_parameters.cell(row=idx, column=1, value="")
    cell_bis = ws_parameters.cell(row=idx, column=1, value="")
#ws_parameters.delete_rows(1)

for row in dataframe_to_rows(df_parameter, index=False, header=True):
    ws_parameters.append(row)
    
# Ajuster la largeur des colonnes en fonction du contenu
for col in ws_parameters.columns:
    max_length = 0
    column = col[0].column  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws_parameters.column_dimensions[get_column_letter(column)].width = adjusted_width

# Put in bold the name line
for row in ws_parameters.iter_rows(min_row=len(text_lines) , max_row=ws_parameters.max_row):#, min_col=1, max_col=1):
    if row[0].value == "Project Input parameters" or row[0].value == "Project Calculated parameters"or row[0].value == "name":# or row[0].value == "name" :
        for cell in row:
            cell.font = Font(bold=True)
    #        cell.fill = PatternFill(start_color="00A1C0", end_color="00A1C0", fill_type="solid")
# Define blue color
fill = PatternFill(start_color='00A1C0', end_color='00A1C0', fill_type='solid')
# Définir le style de police avec la couleur de texte blanche
white_font = Font(color="FFFFFF")

# Save
#loaded_file_name = load_file_name()



# Récupérer le nom de fichier en tant qu'argument de ligne de commande
if len(sys.argv) < 2:
    print("Usage: python annex_c_simapro.py <file_name>")
    sys.exit(1)

file_name_saved = sys.argv[1]


"""
if file_name_saved:
    print("File name is :", file_name_saved)
else:
    print("File name missing.")
    """
#print(file_name_saved)
#print(save_path)
#file_name = "Annex_C.xlsx"
file_save_path = os.path.join(save_path, file_name_saved)

# Enregistrer le classeur Excel
wb.save(file_save_path)
# format_next_cell = False
# """
# # Put in blue the Product line
# for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
#     if any(cell.value == "Products" for cell in row if cell.value is not None):
#         for cell in row:
#             if format_next_cell:
#                 cell.fill = fill
#                 format_next_cell = False  # Reset the flag
#             if cell.value == "Products":
#                 format_next_cell = True  # Set the flag to format the next cell
#             #cell.fill = fill
# """

# In[ ]:




