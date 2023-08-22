#!/usr/bin/env python
# coding: utf-8

# # Ecriture de l'annexe C
# Resources.

# In[1]:


import csv
import pandas as pd
import numpy as np
import pickle
import openpyxl
import sys
import os


# In[ ]:


#File path from the interface
def load_file_path():
    try:
        with open('file_path.pickle', 'rb') as file:
            file_path = pickle.load(file)
            return file_path
    except FileNotFoundError:
        return None

# Utilisation de la fonction pour récupérer le chemin du fichier
file_path = load_file_path()

#Save path 

def load_save_path():
    try:
        with open('save_path.pickle', 'rb') as file:
            save_path = pickle.load(file)
            return save_path
    except FileNotFoundError:
        return None

def load_file_name():
    global file_name_saved  # Ajoutez cette ligne pour indiquer que vous utilisez la variable globale
    try:
        with open('file_name_saved.pickle', 'rb') as file:
            file_name_saved = pickle.load(file)
            if not file_name_saved.endswith('.xlsx'):
                file_name_saved += '.xlsx'
            return file_name_saved
    except FileNotFoundError:
        return None
        
file_name_saved = load_file_name()
#print(file_name_saved)
# Utilisation des fonctions pour récupérer les informations nécessaires
#file_path = load_file_path()
save_path = load_save_path()


# In[2]:


# Specify the path to the .xlsx file you want to open
#excel_file = "C:/Users/u119708/Documents/Untitled Folder/test_bw2.xlsx"

# Load the workbook
workbook = openpyxl.load_workbook(file_path)

# You can now access the sheets of the workbook and cells
active_sheet = workbook.active
data_list=[]
empty_list=[]
test_exchanges = False
i=0
for line in active_sheet.iter_rows(values_only=True):
    # "values_only=True" renvoie les valeurs de chaque cellule sans les objets Cell
    #print(line[0])
    
    #add Exchanges
    i=+1
    if "Exchanges" in line:
        test_exchanges = True
        data_list.append(empty_list)
        #data_list.append(line)
        #print(row)
    elif "Activity" in  line:
        test_exchanges = False
         #print("yes")
        
        #print(line)
    elif test_exchanges:
        # If we are between "Products" and "Avoided products", add the line to the list
        data_list.append(line)
        #print("yes")

        
# Close the workbook when you're done
workbook.close()


# In[3]:


annex_C=pd.DataFrame(data_list)


# In[4]:


# Find index of "products" lines
product_indices = annex_C[annex_C[annex_C.columns[0]] == 'Products'].index
CIRAIG_ecoinvent_index = [index for index, value in enumerate(annex_C.iloc[:, 0]) if str(value).startswith('z.')]
rows_to_remove_1=[]
rows_to_remove_2=[]
i=0
#for n in product_indices:
for i in range (0,len(product_indices)):
    n=product_indices[i]
    #print(n)
    if n + 1 in CIRAIG_ecoinvent_index:
        rows_to_remove_1.append(product_indices[i])
        rows_to_remove_2.append(product_indices[i+1]-1)
rows_to_remove=[]
for start, end in zip(rows_to_remove_1, rows_to_remove_2):
    sublist = list(range(start, end + 1))
    rows_to_remove.extend(sublist)
annex_C.drop(index=rows_to_remove, inplace=True)
annex_C.reset_index(drop=True, inplace=True)
#annex_C


# In[5]:


#Name	unité	quantité	Location	Distribution Type	distrib	distrib	distrib	Comments
#name	amount	database	location	unit	type	reference product
# Colonne 1 reste identique
# Colonne 2 devient la nouvelle 3eme
# Ancienne 3eme est supprimée
# Colonne 4 reste identique
# Colonne 5 devient la 2eme 
#6 biosphere/technosphere
# Autres colonnes sont supprimées

new_order= [0, 4,1,3, 2]  # Spécifiez l'ordre souhaité


# Sélectionnez les colonnes dans le nouvel ordre en utilisant .iloc
annex_C = annex_C.iloc[:, new_order]
#annex_C


# ## Parameters

# In[6]:


#Reading and copy of the relevant elements of the csv.
parameter_list=[]
test_project_parameters=False

i=0
for line in active_sheet.iter_rows(values_only=True):
    # "values_only=True" renvoie les valeurs de chaque cellule sans les objets Cell
    #print(line[0])
    
    #add Exchanges
    i=+1
    if "Project parameters" in line:
        test_project_parameters = True
        parameter_list.append(empty_list)
        parameter_list.append(line)
        #print(row)
    elif "Database" in  line:
        test_project_parameters = False
         #print("yes")
        
        #print(line)
    elif test_project_parameters:
        # If we are between "Products" and "Avoided products", add the line to the list
        parameter_list.append(line)
       
# Close the workbook when you're done
workbook.close()
df_parameter=pd.DataFrame(parameter_list)
#df_parameter


# In[7]:


#new definition of df_parameter
#name	amount	uncertainty type	loc/scale	minimum	maximum	6(aucune idée)	description	
#name	amount	uncertainty type	loc	scale	minimum	maximum	description	group	label
col3_index = 3  # Indice de la colonne 2
col4_index = 4  # Indice de la colonne 3

df_parameter.replace(to_replace=[None, np.nan], value='', inplace=True)

# Convertissez les valeurs des colonnes 2 et 3 en chaînes, puis fusionnez-les dans la colonne 2
df_parameter.iloc[:, col3_index] = df_parameter.iloc[:, col3_index].astype(str) + ' ' + df_parameter.iloc[:, col4_index].astype(str)
#df_parameter.iloc[:, col3_index]= df_parameter.apply(lambda row: ' '.join(filter(None, [row[col3_index], row[col4_index]])), axis=1)

# Supprimez la colonne 3 en utilisant les indices
df_parameter.drop(df_parameter.columns[col4_index], axis=1, inplace=True)
df_parameter.reset_index(drop=True, inplace=True)
new_order_param= [0, 1, 2,3,4,5,7,6]  # Spécifiez l'ordre souhaité

# Sélectionnez les colonnes dans le nouvel ordre en utilisant .iloc
df_parameter = df_parameter.iloc[:, new_order_param]
#df_parameter.columns = ['Name', 'Value', 'Distribution Type', '', '', '', '', 'Description']
#df_parameter.reset_index(inplace=True)
#df_parameter


# In[11]:


from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Creation of new Workbook
wb = Workbook()

# Activate the sheet
ws = wb.active
ws.title = "System description"

# Description project
text_lines = ['Annex C : Projet']#, 'Projet']

# Add lines to Excel file
for idx, text in enumerate(text_lines, start=1):
    cell = ws.cell(row=idx, column=1, value=text)

# Write dataframe and delete first row
for r_idx, row in enumerate(dataframe_to_rows(annex_C, index=False), start=len(text_lines)+1):
    for c_idx, value in enumerate(row, start=1):
        
        #print(value)
        #ws.cell(row=r_idx, column=c_idx, value=value)
        
        try:
            ws.cell(row=r_idx, column=c_idx, value=value)
        except IllegalCharacterError as e:
            # Handle the exception here, e.g., replace the problematic characters or skip the cell
            print(f"Exception: {e}. Cell content: {value}")
ws.delete_rows(2)


# Put in blue the Product line
for row in ws.iter_rows(min_row=len(text_lines) + 2, max_row=ws.max_row, min_col=1, max_col=1):
    if row[0].value == "Products":
        for cell in row:
            cell.font = Font(bold=True)
    #        cell.fill = PatternFill(start_color="00A1C0", end_color="00A1C0", fill_type="solid")
# Define blue color
# Define blue color
fill = PatternFill(start_color='00A1C0', end_color='00A1C0', fill_type='solid')
# Définir le style de police avec la couleur de texte blanche
white_font = Font(color="FFFFFF")

# Put in bold the name line
for row in ws.iter_rows(min_row=len(text_lines) + 2, max_row=ws.max_row):#, min_col=1, max_col=1):
    if row[0].value == "name":
        for cell in row:
            cell.font = Font(bold=True)

# Flag to indicate if the next cell should be formatted

blue_next_row = False

# Parcourir les lignes de la feuille de calcul
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if blue_next_row:
        # Mettre en bleu la ligne courante
        for cell in row:
            cell.fill = fill
            cell.font = white_font
            #print("ok")
        # Réinitialiser le drapeau après avoir mis en bleu la ligne
        blue_next_row = False
    
    if any(cell.value == "name" for cell in row if cell.value is not None):
        # Activer le drapeau pour mettre en bleu la ligne suivante
        blue_next_row = True
        
cell = ws.cell(row=1, column=1)
cell.value = "Annex C : Projet"
cell.font = Font(bold=True, size=16)

import openpyxl

# Définir la largeur des colonnes en pixels
col_width_col1 = 70  # Largeur de la première colonne en pixels
col_width_others = 15  # Largeur des autres colonnes en pixels

# Définir la largeur de la première colonne (indice 1)
ws.column_dimensions[ws.cell(row=1, column=1).column_letter].width = col_width_col1

# Définir la largeur des autres colonnes (à partir de l'indice 2)
for col_idx, col in enumerate(ws.columns, start=1):
    if col_idx != 1:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = col_width_others


#Add the parameter tab

# New Sheet "Parameters"
ws_parameters = wb.create_sheet(title="Parameters")

for idx, text in enumerate(text_lines, start=1):
    cell_bis = ws_parameters.cell(row=idx, column=1, value="")
    cell_bis = ws_parameters.cell(row=idx, column=1, value="")
#ws_parameters.delete_rows(1)

for row in dataframe_to_rows(df_parameter, index=False, header=False):
    ws_parameters.append(row)
    
# Put in bold the name line
for row in ws_parameters.iter_rows(min_row=len(text_lines) + 2, max_row=ws_parameters.max_row):#, min_col=1, max_col=1):
    if row[0].value == "Project parameters" or row[0].value == "name" :
        for cell in row:
            cell.font = Font(bold=True)
# Supprimer la ligne 0
ws_parameters.delete_rows(1)

# Ajuster la largeur des colonnes en fonction du contenu
for col in ws_parameters.columns:
    max_length = 0
    column = col[0].column  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws_parameters.column_dimensions[get_column_letter(column)].width = adjusted_width

# Récupérer le nom de fichier en tant qu'argument de ligne de commande
if len(sys.argv) < 2:
    print("Usage: python annex_c_simapro.py <file_name>")
    sys.exit(1)

file_name_saved = sys.argv[1]


file_save_path = os.path.join(save_path, file_name_saved)

# Enregistrer le classeur Excel
wb.save(file_save_path)
# Save
#wb.save('Test_bw2_excel.xlsx')


# In[ ]:




