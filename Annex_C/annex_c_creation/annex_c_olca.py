#!/usr/bin/env python
# coding: utf-8

# In[ ]:


#Fichier JSON-LD/ZIP
import zipfile
import json
import pandas as pd
import numpy as np
import pickle
import openpyxl
import sys
import os


# In[ ]:


#File path from the interface
def load_file_path():
    try:
        with open('file_path.pickle', 'rb') as file:
            file_path = pickle.load(file)
            return file_path
    except FileNotFoundError:
        return None

# Utilisation de la fonction pour récupérer le chemin du fichier
file_path = load_file_path()

#Save path 

def load_save_path():
    try:
        with open('save_path.pickle', 'rb') as file:
            save_path = pickle.load(file)
            return save_path
    except FileNotFoundError:
        return None

def load_file_name():
    global file_name_saved  # Ajoutez cette ligne pour indiquer que vous utilisez la variable globale
    try:
        with open('file_name_saved.pickle', 'rb') as file:
            file_name_saved = pickle.load(file)
            if not file_name_saved.endswith('.xlsx'):
                file_name_saved += '.xlsx'
            return file_name_saved
    except FileNotFoundError:
        return None
        
file_name_saved = load_file_name()
#print(file_name_saved)
# Utilisation des fonctions pour récupérer les informations nécessaires
#file_path = load_file_path()
save_path = load_save_path()


# In[133]:


######## Take processes
process_list=[]
exchange_list=[]
empty_list=[""]*8
full_annex_c = None 
with zipfile.ZipFile(file_path, 'r') as zip_ref:
    # Obtenir la liste des noms de fichiers dans le ZIP
    liste_fichiers = zip_ref.namelist()
    # Filtrer les fichiers JSON dans le dossier "process"
    fichiers_json_process = [fichier for fichier in liste_fichiers if fichier.startswith('processes/') and fichier.endswith('.json')]
    for fichier_json in fichiers_json_process:
        with zip_ref.open(fichier_json) as json_file:
            json_data = json_file.read()
            # Charger les données JSON
            data = json.loads(json_data.decode('utf-8'))
            process_list = []
            exchange_list = []
            
            # Extraire la valeur du champ "name"
            name_process = data.get("name")
            if "@type" in data and data["@type"] == "Process" and "name" in data:
                name_process = data["name"]
                description_process = data["description"]
                if "location" in data and "name" in data["location"]:
                    location_process=data["location"]["name"]
                else:
                    location_process=""
            #print(name_process)
            exchange_list=[name_process,"","","",description_process]
            #process_list.append(exchange_list)
            if "exchanges" in data:
                exchanges = data["exchanges"]
                for exchange in exchanges:
                    if "description" in exchange :
                        description = exchange["description"]
                    else: 
                        description=""
                    if "flow" in exchange and "refUnit" in exchange["flow"]:
                        refUnit = exchange["flow"]["refUnit"]
                    
                    if "amount" in exchange :
                        amount = exchange["amount"]
                    if "isInput" in exchange :
                        isInput = exchange["isInput"]
                    if "isAvoidedProduct"in exchange :
                        isAvoidedProduct = exchange["isAvoidedProduct"]
                    if "defaultProvider"in exchange and "name" in exchange["defaultProvider"]:
                        flow_name = exchange["defaultProvider"]["name"]
                    else:
                        if "flow" in exchange and "name" in exchange["flow"]:
                            flow_name = exchange["flow"]["name"]
                    if "defaultProvider"in exchange and "location" in exchange["defaultProvider"]:
                        location = exchange["defaultProvider"]["location"]
                    else:
                        location=""
                    if "isQuantitativeReference" in exchange :
                        isQuantitativeReference = exchange["isQuantitativeReference"]
                        flow_name=name_process
                        description=description_process
                        location=location_process
                    else :
                        isQuantitativeReference=False
                    exchange_list=[flow_name,refUnit,amount,location,description,isInput,isAvoidedProduct,isQuantitativeReference]
                    process_list.append(exchange_list)
            
            annex_c=pd.DataFrame(process_list,columns=['name', 'unit','amount','location','description','input','avoided','reference'])
            # Créer un nouveau DataFrame pour les lignes avec 'input' égal à True
            input_true_rows = annex_c[annex_c['input'] == True]
            # Créer un nouveau DataFrame pour les lignes avec 'input' égal à False
            input_false_rows = annex_c[(annex_c['input'] == False) & (annex_c['reference'] == False)]
            #quantitive reference
            reference_true_rows = annex_c[annex_c['reference'] == True]
            # Concaténer les deux DataFrames pour obtenir les lignes avec 'input' égal à True en premier
            sorted_annex_c = pd.concat([reference_true_rows, input_true_rows, input_false_rows])
            # Ajouter sorted_annex_c à full_annex_c
            if full_annex_c is None:
                full_annex_c = sorted_annex_c
            else:
                full_annex_c = pd.concat([full_annex_c, sorted_annex_c])




# In[135]:
empty_list=["","","","","","","",""]
new_rows=[]
column_names=['name', 'unit','value','location ','description','','','']
new_full_annex_c=None
#new_full_annex_c=None
for idx, row in full_annex_c.iterrows():
    #if idx == 0:  # Ignorer la première ligne (index 0)
     #   continue
    if row['reference'] == True:
        #new_rows.append(pd.Series(dtype=str))
        new_rows.append(empty_list) # Ajoutez une ligne vide (série vide
        new_rows.append(column_names) 
    new_rows.append(row)  # Ajoutez la ligne existante
#new_rows
# Créez un nouveau DataFrame en utilisant les lignes préparées
new_full_annex_c = pd.DataFrame(new_rows,columns=column_names)


# In[136]:

new_full_annex_c = new_full_annex_c.drop(new_full_annex_c.columns[-3:], axis=1)
new_full_annex_c = new_full_annex_c.drop(new_full_annex_c.index[0])

#new_full_annex_c


# ## Parameters

# In[137]:


with zipfile.ZipFile(file_path, 'r') as zip_ref:
    liste_fichiers = zip_ref.namelist()
    fichiers_json_process = [fichier for fichier in liste_fichiers if fichier.startswith('parameters/') and fichier.endswith('.json')]
    final_list=[]
    for fichier_json in fichiers_json_process:
        with zip_ref.open(fichier_json) as json_file:
            json_data = json_file.read()
            # Charger les données JSON
            data = json.loads(json_data.decode('utf-8'))
            parameter_list = []
            #final_list = []
            
            #process_list.append(exchange_list)
            if "name" in data:
                name_parameter = data["name"]
            if "isInputParameter" in data:
                isInputParameter = data["isInputParameter"]
            if "description" in data:
                description = data["description"]
            if "value" in data:
                amount = data["value"]
            else:
                description=""
            if "uncertainty" in data:
                distributionType=data["uncertainty"]["distributionType"]
                if "minimum" in data["uncertainty"]:
                    parameter_min=data["uncertainty"]["minimum"]
                if "mode" in data["uncertainty"]:
                    parameter_mode=data["uncertainty"]["mode"]
                if "maximum" in data["uncertainty"]:
                    parameter_max=data["uncertainty"]["maximum"]
                if "geomMean"in data["uncertainty"]:
                    parameter_geomMean=data["uncertainty"]["geomMean"]
                if "geomSd"in data["uncertainty"]:
                    parameter_geomSd=data["uncertainty"]["geomSd"]
                if "mean"in data["uncertainty"]:
                    parameter_mean=data["uncertainty"]["mean"]
                if "sd"in data["uncertainty"]:
                    parameter_sd=data["uncertainty"]["sd"]
                if distributionType=="LOG_NORMAL_DISTRIBUTION":
                    parameter_list=[name_parameter,amount,distributionType,f"geomMean={parameter_geomMean}",f"geomSd={parameter_geomSd}","",description,isInputParameter]
                    #final_list.append(parameter_list)
                    #print(final_list)
                if distributionType=="TRIANGLE_DISTRIBUTION":
                    parameter_list=[name_parameter,amount,distributionType,f"min={parameter_min}",f"mode={parameter_mode}",f"max={parameter_max}",description,isInputParameter]
                    #final_list.append(parameter_list)
                    #print(final_list)
                if distributionType=="NORMAL_DISTRIBUTION":
                    parameter_list=[name_parameter,amount,distributionType,f"mean={parameter_mean}",f"sd={parameter_sd}","",description,isInputParameter]
                    #final_list.append(parameter_list)
                if distributionType=="UNIFORM_DISTRIBUTION":
                    parameter_list=[name_parameter,amount,distributionType,f"min={parameter_min}",f"max={parameter_max}","",description,isInputParameter]
                    #final_list.append(parameter_list)
        
            else:
                parameter_list=[name_parameter,amount,"Undefined","","","",description,isInputParameter]
        final_list.append(parameter_list)
            #df_parameter=pd.DataFrame(parameter_list,columns=['name', 'amount','distribution type','','','','description','isInputparameter'])
            
#len(final_list)

df_parameter=pd.DataFrame(final_list,columns=['name', 'amount','distribution type','','','','description','isInputparameter'])
            
#df_parameter        
                    


# In[138]:


import pandas as pd
import numpy as np
process_list=[]
exchange_list=[]
empty_list=[""]*8
full_annex_c = None 
#Ajouter les input parameters
with zipfile.ZipFile(file_path, 'r') as zip_ref:
    # Obtenir la liste des noms de fichiers dans le ZIP
    liste_fichiers = zip_ref.namelist()
    final_list_2 = []
    # Filtrer les fichiers JSON dans le dossier "process"
    fichiers_json_process = [fichier for fichier in liste_fichiers if fichier.startswith('processes/') and fichier.endswith('.json')]
    for fichier_json in fichiers_json_process:
        with zip_ref.open(fichier_json) as json_file:
            json_data = json_file.read()
            data = json.loads(json_data.decode('utf-8'))
            param_list_2 = []
            if "parameters" in data :
                parameters=data["parameters"]
                for parameter in parameters:
                    name_param=parameter["name"]
                    if "name" in parameter:
                        name_parameter = parameter["name"]
                    if "value" in parameter:
                        amount = parameter["value"]
                    if "isInputParameter" in parameter:
                        isInputParameter = parameter["isInputParameter"]
                        if isInputParameter==False:
                            
                            amount=parameter["formula"]
                    if "description" in parameter:
                        description = parameter["description"]
                    else:
                        description=""
                    if "uncertainty" in parameter:
                        distributionType=parameter["uncertainty"]["distributionType"]
                        if "minimum" in parameter["uncertainty"]:
                            parameter_min=parameter["uncertainty"]["minimum"]
                        if "mode" in parameter["uncertainty"]:
                            parameter_mode=parameter["uncertainty"]["mode"]
                        if "maximum" in parameter["uncertainty"]:
                            parameter_max=parameter["uncertainty"]["maximum"]
                        if "geomMean"in parameter["uncertainty"]:
                            parameter_geomMean=parameter["uncertainty"]["geomMean"]
                        if "geomSd"in parameter["uncertainty"]:
                            parameter_geomSd=parameter["uncertainty"]["geomSd"]
                        if "mean"in parameter["uncertainty"]:
                            parameter_mean=parameter["uncertainty"]["mean"]
                        if "sd"in parameter["uncertainty"]:
                            parameter_sd=parameter["uncertainty"]["sd"]
                        if distributionType=="LOG_NORMAL_DISTRIBUTION":
                            param_list_2=[name_parameter,amount,distributionType,f"geomMean={parameter_geomMean}",f"geomSd={parameter_geomSd}","",description,isInputParameter]
                        if distributionType=="TRIANGLE_DISTRIBUTION":
                            param_list_2=[name_parameter,amount,distributionType,f"min={parameter_min}",f"mode={parameter_mode}",f"max={parameter_max}",description,isInputParameter]
                        if distributionType=="NORMAL_DISTRIBUTION":
                            param_list_2=[name_parameter,amount,distributionType,f"mean={parameter_mean}",f"sd={parameter_sd}","",description,isInputParameter]
                        if distributionType=="UNIFORM_DISTRIBUTION":
                            param_list_2=[name_parameter,amount,distributionType,f"min={parameter_min}",f"max={parameter_max}","",description,isInputParameter]

                    else:
                        param_list_2=[name_parameter,amount,"Undefined","","","",description,isInputParameter]
                    final_list_2.append(param_list_2)
#final_list_2
df_parameter_2=pd.DataFrame(final_list_2,columns=['name', 'amount','distribution type','','','','description','isInputparameter'])
            
#df_parameter_2                  


# In[139]:


df_parameters_final=pd.concat([df_parameter, df_parameter_2], ignore_index=True)
df_parameters_final.drop(df_parameters_final.columns[-1], axis=1, inplace=True)
#df_parameters_final

column_names=['name', 'value', 'distribution type', '', '', '', 'description']

df_parameters_final.columns = column_names

df_parameters_final_2 = df_parameters_final.drop_duplicates()

# In[ ]:


from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Creation of new Workbook
wb = Workbook()

# Activate the sheet
ws = wb.active
ws.title = "System description"

# Description project
text_lines = ['Annex C : Projet']#, 'Projet']

# Add lines to Excel file
for idx, text in enumerate(text_lines, start=1):
    cell = ws.cell(row=idx, column=1, value=text)

# Write dataframe and delete first row
for r_idx, row in enumerate(dataframe_to_rows(new_full_annex_c, index=False), start=len(text_lines)+1):
    for c_idx, value in enumerate(row, start=1):
        
        #print(value)
        #ws.cell(row=r_idx, column=c_idx, value=value)
        
        try:
            ws.cell(row=r_idx, column=c_idx, value=value)
        except IllegalCharacterError as e:
            # Handle the exception here, e.g., replace the problematic characters or skip the cell
            print(f"Exception: {e}. Cell content: {value}")
ws.delete_rows(2)


# Put in blue the Product line
for row in ws.iter_rows(min_row=len(text_lines) + 2, max_row=ws.max_row, min_col=1, max_col=1):
    if row[0].value == "Products":
        for cell in row:
            cell.font = Font(bold=True)
    #        cell.fill = PatternFill(start_color="00A1C0", end_color="00A1C0", fill_type="solid")
# Define blue color
# Define blue color
fill = PatternFill(start_color='00A1C0', end_color='00A1C0', fill_type='solid')
# Définir le style de police avec la couleur de texte blanche
white_font = Font(color="FFFFFF")

# Put in bold the name line
for row in ws.iter_rows(min_row=len(text_lines) , max_row=ws.max_row):#, min_col=1, max_col=1):
    if row[0].value == "name":
        for cell in row:
            cell.font = Font(bold=True)

# Flag to indicate if the next cell should be formatted

blue_next_row = False

# Parcourir les lignes de la feuille de calcul
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if blue_next_row:
        # Mettre en bleu la ligne courante
        for cell in row:
            cell.fill = fill
            cell.font = white_font
            #print("ok")
        # Réinitialiser le drapeau après avoir mis en bleu la ligne
        blue_next_row = False
    
    if any(cell.value == "name" for cell in row if cell.value is not None):
        # Activer le drapeau pour mettre en bleu la ligne suivante
        blue_next_row = True
        
cell = ws.cell(row=1, column=1)
cell.value = "Annex C : Projet"
cell.font = Font(bold=True, size=16)

import openpyxl

# Définir la largeur des colonnes en pixels
col_width_col1 = 70  # Largeur de la première colonne en pixels
col_width_others = 15  # Largeur des autres colonnes en pixels

# Définir la largeur de la première colonne (indice 1)
ws.column_dimensions[ws.cell(row=1, column=1).column_letter].width = col_width_col1

# Définir la largeur des autres colonnes (à partir de l'indice 2)
for col_idx, col in enumerate(ws.columns, start=1):
    if col_idx != 1:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = col_width_others


#Add the parameter tab

# New Sheet "Parameters"
ws_parameters = wb.create_sheet(title="Parameters")

for idx, text in enumerate(text_lines, start=1):
    cell_bis = ws_parameters.cell(row=idx, column=1, value="")
    cell_bis = ws_parameters.cell(row=idx, column=1, value="")
#ws_parameters.delete_rows(1)

for row in dataframe_to_rows(df_parameters_final_2, index=False, header=True):
    ws_parameters.append(row)
    
# Put in bold the name line
for row in ws_parameters.iter_rows(min_row=len(text_lines) , max_row=ws_parameters.max_row):#, min_col=1, max_col=1):
    if row[0].value == "Project parameters" or row[0].value == "name" :
        for cell in row:
            cell.font = Font(bold=True)
# Supprimer la ligne 0
ws_parameters.delete_rows(1)

# Ajuster la largeur des colonnes en fonction du contenu
for col in ws_parameters.columns:
    max_length = 0
    column = col[0].column  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws_parameters.column_dimensions[get_column_letter(column)].width = adjusted_width

# Récupérer le nom de fichier en tant qu'argument de ligne de commande
if len(sys.argv) < 2:
    print("Usage: python annex_c_olca.py <file_name>")
    sys.exit(1)

file_name_saved = sys.argv[1]


file_save_path = os.path.join(save_path, file_name_saved)

# Enregistrer le classeur Excel
wb.save(file_save_path)


# In[ ]:





# In[ ]:




